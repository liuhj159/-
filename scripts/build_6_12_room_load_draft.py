from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "6-12_plan_scan"


EXCLUDE_LABELS = {"送风井", "铝合金百叶", "水吧台", "企业文化墙", "投影墙", "鱼缸", "冰箱"}


def classify(name: str) -> tuple[str, int]:
    if any(key in name for key in ["洽谈室", "会议室"]):
        return "会议/洽谈", 220
    if "接待" in name:
        return "接待/休闲", 200
    if any(key in name for key in ["造价部", "采购部", "招标部", "综合部", "财务部", "经理", "主管"]):
        return "办公", 180
    if any(key in name for key in ["档案室", "储藏间", "打印室"]):
        return "辅助用房", 120
    if any(key in name for key in ["包厢", "雪茄室", "茶道"]):
        return "包厢/会客", 220
    if "清洗间" in name:
        return "辅助/清洗", 150
    return "待确认", 180


def load_room_labels() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with (OUT_DIR / "room_name_candidates.csv").open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            name = row["text"].strip()
            if name in EXCLUDE_LABELS:
                continue
            rows.append(row)
    return rows


def build_rows(room_labels: list[dict[str, str]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for index, row in enumerate(room_labels, start=1):
        name = row["text"].strip()
        function, unit_load = classify(name)
        rows.append(
            {
                "序号": index,
                "房间名称": name,
                "房间功能建议": function,
                "面积m2": "",
                "单位冷负荷W/m2建议": unit_load,
                "修正系数": 1.0,
                "房间冷负荷W": "",
                "面积来源": "未从图纸识别到面积标注；自动面域化未形成可靠房间边界",
                "计算公式": "面积m2 × 单位冷负荷W/m2 × 修正系数",
                "图层": row["layer"],
                "文字坐标X": row["x"],
                "文字坐标Y": row["y"],
                "状态": "待确认面积后计算",
                "备注": "单位负荷为经验建议，需结合城市气象、围护结构、朝向、新风和人员密度由设计师确认",
            }
        )
    return rows


def append_unnamed_furniture_rooms(rows: list[dict[str, object]]) -> None:
    path = OUT_DIR / "unnamed_furniture_room_candidates.csv"
    if not path.exists():
        return
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            room_no = row.get("temp_room_no", "").strip()
            if not room_no:
                continue
            rows.append(
                {
                    "序号": len(rows) + 1,
                    "房间名称": room_no,
                    "房间功能建议": "有办公家具未命名",
                    "面积m2": "",
                    "单位冷负荷W/m2建议": 180,
                    "修正系数": 1.0,
                    "房间冷负荷W": "",
                    "面积来源": f"家具聚类候选，家具点包围盒约 {row.get('bbox_area_m2', '')} m2；不是房间净面积",
                    "计算公式": "面积m2 × 单位冷负荷W/m2 × 修正系数",
                    "图层": row.get("layers", ""),
                    "文字坐标X": row.get("center_x", ""),
                    "文字坐标Y": row.get("center_y", ""),
                    "状态": "未命名房间，待确认名称和面积后计算",
                    "备注": "由办公家具/活动家具聚类识别；临时编号需人工确认是否为空调服务区域",
                }
            )


def write_csv(rows: list[dict[str, object]], headers: list[str]) -> Path:
    path = OUT_DIR / "room_load_draft.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    return path


def write_xlsx(rows: list[dict[str, object]], headers: list[str]) -> Path:
    path = OUT_DIR / "6-12房间负荷计算底稿.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "房间负荷底稿"
    ws.append(headers)
    for row in rows:
        ws.append([row[header] for header in headers])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        1: 8,
        2: 18,
        3: 16,
        4: 12,
        5: 20,
        6: 10,
        7: 14,
        8: 38,
        9: 26,
        10: 16,
        11: 14,
        12: 14,
        13: 16,
        14: 48,
    }
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"

    issue = wb.create_sheet("待确认事项")
    issue.append(["问题编号", "问题", "影响", "建议处理"])
    issue.append(
        [
            "Q-001",
            "图纸未识别到房间面积标注，且墙体线未能自动闭合成可靠房间边界",
            "无法自动计算确定房间总冷负荷",
            "提供带面积房间多段线、天正面积表、原 CAD 面积统计，或由设计师确认面积后回填",
        ]
    )
    issue.append(
        [
            "Q-002",
            "项目城市、围护结构、朝向、窗墙比、新风和人员密度未提供",
            "单位冷负荷只能作为经验建议",
            "补充项目所在地和设计标准后复核单位负荷",
        ]
    )
    for cell in issue[1]:
        cell.font = Font(bold=True)
    issue.column_dimensions["A"].width = 12
    issue.column_dimensions["B"].width = 52
    issue.column_dimensions["C"].width = 36
    issue.column_dimensions["D"].width = 58

    wb.save(path)
    return path


def write_report(rows: list[dict[str, object]]) -> Path:
    path = OUT_DIR / "6-12_plan_scan_report.md"
    lines = [
        "# 6-12平面图扫描报告",
        "",
        "## 输入资料",
        "",
        "- DWF：`6-12平面.dwf`",
        "- 原始 DWG：`F:\\xwechat_files\\liuguoqiang304127_6d46\\msg\\file\\2026-06\\6-12平面.dwg`",
        "- 工作副本：`outputs/6-12_plan_scan/src/6-12_plan.dwg`",
        "- 派生 DXF：`outputs/6-12_plan_scan/6_12_plan_R2018.dxf`",
        "- 原始资料未修改。",
        "",
        "## 抽取结果",
        "",
        "- 识别 TEXT/MTEXT：76 条。",
        f"- 识别到可作为房间/功能区名称的标注：{len(rows)} 条。",
        "- 未识别到可靠房间面积文字标注。",
        "- 闭合多段线主要来自家具图层，不能作为房间面积依据。",
        "- 使用墙/窗/柱线做自动面域化后，未能形成可包围房间文字的可靠房间边界。",
        "",
        "## 已识别房间/区域",
        "",
    ]
    lines.extend(f"- {row['房间名称']}：{row['房间功能建议']}，建议单位冷负荷 {row['单位冷负荷W/m2建议']} W/m2" for row in rows)
    lines.extend(
        [
            "",
            "## 结论",
            "",
            "本次可以可靠提取房间/区域名称，但不能从当前图纸自动得到可信面积。因此负荷底稿已先生成单位冷负荷建议，房间总冷负荷需待面积确认后计算。",
            "",
            "## 成果文件",
            "",
            "- `outputs/6-12_plan_scan/room_load_draft.csv`",
            "- `outputs/6-12_plan_scan/6-12房间负荷计算底稿.xlsx`",
            "- `outputs/6-12_plan_scan/text_entities.csv`",
            "- `outputs/6-12_plan_scan/room_name_candidates.csv`",
            "- `outputs/6-12_plan_scan/room_area_polygonized_candidates.csv`",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    return path


def main() -> None:
    room_labels = load_room_labels()
    rows = build_rows(room_labels)
    append_unnamed_furniture_rooms(rows)
    headers = list(rows[0].keys()) if rows else []
    csv_path = write_csv(rows, headers)
    xlsx_path = write_xlsx(rows, headers)
    report_path = write_report(rows)
    print(f"rooms={len(rows)}")
    print(f"csv={csv_path}")
    print(f"xlsx={xlsx_path}")
    print(f"report={report_path}")


if __name__ == "__main__":
    main()
