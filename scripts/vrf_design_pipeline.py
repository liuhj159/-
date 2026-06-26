from __future__ import annotations

import argparse
import csv
import hashlib
import math
import re
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

import ezdxf
from ezdxf.enums import TextEntityAlignment
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_ROOT = ROOT / "outputs"

TEXT_TYPES = {"TEXT", "MTEXT"}
POLY_TYPES = {"LWPOLYLINE", "POLYLINE"}

EXCLUDE_TEXT_PATTERNS = [
    r"^A\d+$",
    r"^\d+[:：]\d+$",
    r"^[\-+~—_]+$",
    r"^DN\d+",
    r"^φ?\d+(\.\d+)?$",
]

AREA_TEXT_RE = re.compile(r"(?P<area>\d+(?:\.\d+)?)\s*(?:m2|㎡|平方米|M2)", re.IGNORECASE)


@dataclass
class SourceRow:
    source_id: str
    file_or_link: str
    type: str
    version_date: str
    owner_role: str
    status: str
    notes: str
    sha256: str


@dataclass
class TextRow:
    source_id: str
    text: str
    layer: str
    x: float
    y: float
    height: float
    evidence: str


@dataclass
class RoomRow:
    room_id: str
    source_id: str
    room_name: str
    function: str
    area_m2: float | None
    area_source: str
    unit_load_w_m2: int
    correction_factor: float
    cooling_load_w: float | None
    cooling_load_kw: float | None
    boundary_layer: str
    center_x: float | None
    center_y: float | None
    status: str
    evidence: str
    assumptions: str


@dataclass
class SelectionRow:
    room_id: str
    room_name: str
    function: str
    cooling_load_kw: float | None
    selected_capacity_kw: float | None
    quantity: int
    indoor_unit_form: str
    candidate_model: str
    capacity_margin_pct: float | None
    conclusion: str
    evidence: str


@dataclass
class IssueRow:
    issue_id: str
    severity: str
    source: str
    description: str
    impact: str
    evidence: str
    owner: str
    recommendation: str
    status: str
    release_conclusion: str


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\\P", " ")).strip()


def entity_text(entity) -> str:
    if entity.dxftype() == "TEXT":
        return clean_text(entity.dxf.text)
    if entity.dxftype() == "MTEXT":
        return clean_text(entity.text)
    return ""


def entity_point(entity) -> tuple[float, float]:
    if entity.dxftype() == "TEXT":
        p = entity.dxf.insert
    else:
        p = entity.dxf.insert
    return float(p.x), float(p.y)


def is_probable_room_name(text: str) -> bool:
    if not text or len(text) > 40:
        return False
    if AREA_TEXT_RE.search(text):
        return False
    if any(re.search(pattern, text, re.IGNORECASE) for pattern in EXCLUDE_TEXT_PATTERNS):
        return False
    room_keywords = [
        "室",
        "厅",
        "房",
        "间",
        "区",
        "办公室",
        "会议",
        "接待",
        "餐",
        "包厢",
        "厨房",
        "茶",
        "财务",
        "经理",
        "前台",
    ]
    return any(keyword in text for keyword in room_keywords)


def classify_room(name: str) -> tuple[str, int]:
    rules = [
        (["会议", "洽谈"], "会议/洽谈", 220),
        (["餐", "包厢", "宴会", "大厅"], "餐饮/包厢", 240),
        (["厨房", "后厨"], "厨房/后厨", 300),
        (["接待", "前台", "休闲"], "接待/休闲", 200),
        (["机房", "弱电", "设备"], "设备/机房", 300),
        (["办公室", "办公", "财务", "经理", "主管", "采购", "造价", "招标"], "办公", 180),
        (["档案", "储藏", "库房", "打印"], "辅助用房", 120),
        (["卫生", "清洗", "更衣"], "辅助/清洗", 150),
    ]
    for keywords, function, unit_load in rules:
        if any(keyword in name for keyword in keywords):
            return function, unit_load
    return "待确认", 180


def select_unit_form(function: str, capacity_kw: float | None) -> str:
    if function in {"会议/洽谈", "餐饮/包厢", "接待/休闲"}:
        return "四面出风嵌入式或薄型风管机候选"
    if function == "厨房/后厨":
        return "高静压风管机候选，需复核油烟和新风排风条件"
    if function == "设备/机房":
        return "专用空调或风管机候选，需按显热和全年运行复核"
    if capacity_kw and capacity_kw > 9:
        return "风管机或多台嵌入式候选"
    return "薄型风管机候选"


def select_capacity(load_kw: float | None) -> tuple[float | None, int, float | None]:
    if load_kw is None or load_kw <= 0:
        return None, 0, None
    capacities = [2.2, 2.8, 3.6, 4.5, 5.6, 7.1, 8.0, 9.0, 11.2, 14.0, 16.0]
    for capacity in capacities:
        if capacity >= load_kw:
            return capacity, 1, (capacity / load_kw - 1) * 100
    qty = math.ceil(load_kw / 14.0)
    capacity = 14.0
    total = qty * capacity
    return capacity, qty, (total / load_kw - 1) * 100


def poly_points(entity) -> list[tuple[float, float]]:
    if entity.dxftype() == "LWPOLYLINE":
        return [(float(p[0]), float(p[1])) for p in entity.get_points("xy")]
    if entity.dxftype() == "POLYLINE":
        return [(float(v.dxf.location.x), float(v.dxf.location.y)) for v in entity.vertices]
    return []


def polygon_area(points: list[tuple[float, float]]) -> float:
    if len(points) < 3:
        return 0.0
    total = 0.0
    for i, (x1, y1) in enumerate(points):
        x2, y2 = points[(i + 1) % len(points)]
        total += x1 * y2 - x2 * y1
    return abs(total) / 2.0


def polygon_centroid(points: list[tuple[float, float]]) -> tuple[float, float]:
    if not points:
        return 0.0, 0.0
    area2 = 0.0
    cx = 0.0
    cy = 0.0
    for i, (x1, y1) in enumerate(points):
        x2, y2 = points[(i + 1) % len(points)]
        cross = x1 * y2 - x2 * y1
        area2 += cross
        cx += (x1 + x2) * cross
        cy += (y1 + y2) * cross
    if abs(area2) < 1e-9:
        return sum(x for x, _ in points) / len(points), sum(y for _, y in points) / len(points)
    return cx / (3 * area2), cy / (3 * area2)


def point_in_polygon(x: float, y: float, points: list[tuple[float, float]]) -> bool:
    inside = False
    j = len(points) - 1
    for i, (xi, yi) in enumerate(points):
        xj, yj = points[j]
        intersects = (yi > y) != (yj > y) and x < (xj - xi) * (y - yi) / ((yj - yi) or 1e-12) + xi
        if intersects:
            inside = not inside
        j = i
    return inside


def read_dxf(path: Path):
    return ezdxf.readfile(path)


def extract_texts(doc, source_id: str) -> list[TextRow]:
    rows: list[TextRow] = []
    for entity in doc.modelspace():
        if entity.dxftype() not in TEXT_TYPES:
            continue
        text = entity_text(entity)
        if not text:
            continue
        x, y = entity_point(entity)
        rows.append(
            TextRow(
                source_id=source_id,
                text=text,
                layer=entity.dxf.layer,
                x=round(x, 3),
                y=round(y, 3),
                height=round(float(getattr(entity.dxf, "height", 0) or 0), 3),
                evidence=f"{source_id}:{entity.dxftype()}:{entity.dxf.layer}@({x:.3f},{y:.3f})",
            )
        )
    return rows


def extract_room_polygons(doc, min_area: float, max_area: float) -> list[dict[str, object]]:
    polygons: list[dict[str, object]] = []
    for entity in doc.modelspace():
        if entity.dxftype() not in POLY_TYPES:
            continue
        is_closed = bool(getattr(entity, "closed", False)) or bool(getattr(entity.dxf, "flags", 0) & 1)
        if not is_closed:
            continue
        points = poly_points(entity)
        raw_area = polygon_area(points)
        if raw_area < min_area or raw_area > max_area:
            continue
        cx, cy = polygon_centroid(points)
        polygons.append(
            {
                "points": points,
                "raw_area": raw_area,
                "layer": entity.dxf.layer,
                "center_x": cx,
                "center_y": cy,
                "handle": entity.dxf.handle,
            }
        )
    return polygons


def nearest_area_text(x: float, y: float, texts: Iterable[TextRow], max_distance: float) -> tuple[float | None, str]:
    best: tuple[float, float, str] | None = None
    for row in texts:
        match = AREA_TEXT_RE.search(row.text)
        if not match:
            continue
        distance = math.hypot(row.x - x, row.y - y)
        if distance > max_distance:
            continue
        area = float(match.group("area"))
        if best is None or distance < best[0]:
            best = (distance, area, row.evidence)
    if best:
        return best[1], best[2]
    return None, ""


def build_rooms(
    source_id: str,
    texts: list[TextRow],
    polygons: list[dict[str, object]],
    unit_scale: float,
    correction_factor: float,
) -> list[RoomRow]:
    room_texts = [row for row in texts if is_probable_room_name(row.text)]
    used_text_ids: set[int] = set()
    rooms: list[RoomRow] = []

    for polygon in sorted(polygons, key=lambda item: float(item["raw_area"])):
        contained = [
            row
            for row in room_texts
            if id(row) not in used_text_ids and point_in_polygon(row.x, row.y, polygon["points"])  # type: ignore[arg-type]
        ]
        if not contained:
            continue
        label = min(
            contained,
            key=lambda row: math.hypot(row.x - float(polygon["center_x"]), row.y - float(polygon["center_y"])),
        )
        used_text_ids.add(id(label))
        area_m2 = float(polygon["raw_area"]) * unit_scale * unit_scale
        function, unit_load = classify_room(label.text)
        load_w = area_m2 * unit_load * correction_factor
        room_no = len(rooms) + 1
        rooms.append(
            RoomRow(
                room_id=f"R-{room_no:03d}",
                source_id=source_id,
                room_name=label.text,
                function=function,
                area_m2=round(area_m2, 2),
                area_source=f"闭合多段线面积，图层 {polygon['layer']}，handle {polygon['handle']}，单位比例 {unit_scale}",
                unit_load_w_m2=unit_load,
                correction_factor=correction_factor,
                cooling_load_w=round(load_w, 1),
                cooling_load_kw=round(load_w / 1000, 2),
                boundary_layer=str(polygon["layer"]),
                center_x=round(float(polygon["center_x"]), 3),
                center_y=round(float(polygon["center_y"]), 3),
                status="自动识别，待设计师复核",
                evidence=f"{label.evidence}; polygon_handle={polygon['handle']}",
                assumptions="单位冷负荷为经验建议，需结合城市气象、围护结构、朝向、新风和人员密度确认",
            )
        )

    for label in room_texts:
        if id(label) in used_text_ids:
            continue
        area_text, area_evidence = nearest_area_text(label.x, label.y, texts, max_distance=10000)
        function, unit_load = classify_room(label.text)
        load_w = area_text * unit_load * correction_factor if area_text else None
        room_no = len(rooms) + 1
        rooms.append(
            RoomRow(
                room_id=f"R-{room_no:03d}",
                source_id=source_id,
                room_name=label.text,
                function=function,
                area_m2=round(area_text, 2) if area_text else None,
                area_source=f"附近面积文字：{area_evidence}" if area_text else "未识别到闭合房间边界或面积文字",
                unit_load_w_m2=unit_load,
                correction_factor=correction_factor,
                cooling_load_w=round(load_w, 1) if load_w else None,
                cooling_load_kw=round(load_w / 1000, 2) if load_w else None,
                boundary_layer="",
                center_x=label.x,
                center_y=label.y,
                status="待确认面积后计算" if not area_text else "面积来自文字候选，待复核",
                evidence=label.evidence,
                assumptions="单位冷负荷为经验建议，需结合城市气象、围护结构、朝向、新风和人员密度确认",
            )
        )
    return rooms


def build_selections(rooms: list[RoomRow], brand: str, series: str) -> list[SelectionRow]:
    rows: list[SelectionRow] = []
    for room in rooms:
        capacity, quantity, margin = select_capacity(room.cooling_load_kw)
        form = select_unit_form(room.function, capacity)
        if capacity:
            model = f"{brand}-{series}-IDU-{int(round(capacity * 10)):03d}"
            conclusion = "候选满足负荷，待设计师按吊顶、噪声、静压、检修和厂家样册确认"
        else:
            model = ""
            conclusion = "缺少面积或负荷，暂不能选型"
        rows.append(
            SelectionRow(
                room_id=room.room_id,
                room_name=room.room_name,
                function=room.function,
                cooling_load_kw=room.cooling_load_kw,
                selected_capacity_kw=capacity,
                quantity=quantity,
                indoor_unit_form=form,
                candidate_model=model,
                capacity_margin_pct=round(margin, 1) if margin is not None else None,
                conclusion=conclusion,
                evidence=f"room={room.room_id}; load={room.cooling_load_kw}kW; brand={brand}; series={series}",
            )
        )
    return rows


def build_issues(inputs: list[Path], rooms: list[RoomRow], brand: str, city: str) -> list[IssueRow]:
    issues: list[IssueRow] = []
    unsupported = [path for path in inputs if path.suffix.lower() in {".dwf", ".dwfx", ".pdf"}]
    if unsupported:
        issues.append(
            IssueRow(
                issue_id=f"QA-{len(issues)+1:03d}",
                severity="S2",
                source="automation",
                description="存在当前脚本不能直接解析的图纸格式",
                impact="DWF/DWFX/PDF 需先转 DWG/DXF 或 OCR 后才能可靠提取房间边界和面积",
                evidence="; ".join(str(path) for path in unsupported),
                owner="automation-engineer",
                recommendation="提供原始 DWG/DXF，或允许按 PDF/OCR 路线生成待复核底稿",
                status="open",
                release_conclusion="conditional pass",
            )
        )
    if any(room.area_m2 is None for room in rooms):
        issues.append(
            IssueRow(
                issue_id=f"QA-{len(issues)+1:03d}",
                severity="S2",
                source="design",
                description="部分房间未识别到可靠面积",
                impact="无法形成确定冷负荷和设备容量",
                evidence="; ".join(room.room_id for room in rooms if room.area_m2 is None),
                owner="designer",
                recommendation="补充带闭合房间边界的 CAD、面积表，或由设计师回填面积",
                status="open",
                release_conclusion="conditional pass",
            )
        )
    if not city:
        issues.append(
            IssueRow(
                issue_id=f"QA-{len(issues)+1:03d}",
                severity="S3",
                source="design",
                description="项目城市未提供",
                impact="室外计算参数、负荷指标和修正系数未能按城市气象复核",
                evidence="命令行参数 --city 为空",
                owner="designer",
                recommendation="补充项目所在地，并引用对应气象参数来源",
                status="open",
                release_conclusion="conditional pass",
            )
        )
    if brand == "通用":
        issues.append(
            IssueRow(
                issue_id=f"QA-{len(issues)+1:03d}",
                severity="S3",
                source="design",
                description="品牌/系列未明确",
                impact="室内机型号、容量余量、配比、管长、分歧管和冷媒追加不能按厂家规则定稿",
                evidence="命令行参数 --brand 使用默认通用",
                owner="designer",
                recommendation="指定品牌和系列，并回查厂家产品样册或设计选型手册",
                status="open",
                release_conclusion="conditional pass",
            )
        )
    return issues


def write_csv(path: Path, rows: list[object]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(asdict(rows[0]).keys()))
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def append_sheet(wb: Workbook, title: str, rows: list[object]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["无记录"])
        return
    headers = list(asdict(rows[0]).keys())
    ws.append(headers)
    for row in rows:
        ws.append([asdict(row).get(header) for header in headers])
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, header in enumerate(headers, start=1):
        width = min(max(len(header) + 4, 12), 38)
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"


def write_workbook(path: Path, sheets: dict[str, list[object]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        append_sheet(wb, title, rows)
    wb.save(path)


def add_layer(doc, name: str, color: int) -> None:
    if name not in doc.layers:
        doc.layers.add(name, color=color)


def write_overlay_dxf(path: Path, rooms: list[RoomRow], selections: list[SelectionRow]) -> None:
    doc = ezdxf.new("R2018")
    msp = doc.modelspace()
    add_layer(doc, "VRF_ROOM_CENTER", 3)
    add_layer(doc, "VRF_INDOOR_UNIT_DRAFT", 1)
    add_layer(doc, "VRF_LOAD_NOTE", 2)
    selection_by_room = {row.room_id: row for row in selections}
    for room in rooms:
        if room.center_x is None or room.center_y is None:
            continue
        x = room.center_x
        y = room.center_y
        size = 450
        msp.add_circle((x, y), radius=size / 2, dxfattribs={"layer": "VRF_ROOM_CENTER", "color": 3})
        msp.add_lwpolyline(
            [(x - size, y - size), (x + size, y - size), (x + size, y + size), (x - size, y + size)],
            close=True,
            dxfattribs={"layer": "VRF_INDOOR_UNIT_DRAFT", "color": 1},
        )
        sel = selection_by_room.get(room.room_id)
        label = f"{room.room_id} {room.room_name}"
        if sel and sel.selected_capacity_kw:
            label += f" {sel.quantity}x{sel.selected_capacity_kw}kW"
        msp.add_text(
            label,
            dxfattribs={"layer": "VRF_LOAD_NOTE", "height": 220, "color": 2},
        ).set_placement((x, y + size + 250), align=TextEntityAlignment.MIDDLE_CENTER)
    doc.saveas(path)


def write_report(
    path: Path,
    source_rows: list[SourceRow],
    rooms: list[RoomRow],
    selections: list[SelectionRow],
    issues: list[IssueRow],
    brand: str,
    series: str,
    city: str,
) -> None:
    calculated = [room for room in rooms if room.cooling_load_kw is not None]
    total_load = sum(room.cooling_load_kw or 0 for room in calculated)
    lines = [
        "# VRF 设计报价自动化底稿",
        "",
        "## 输入资料",
        "",
    ]
    lines.extend(f"- {row.source_id}: `{row.file_or_link}`，类型 {row.type}，状态 {row.status}" for row in source_rows)
    lines.extend(
        [
            "",
            "## 本次口径",
            "",
            f"- 城市：{city or '未提供，待确认'}",
            f"- 品牌/系列：{brand} / {series}",
            "- 面积单位：默认 CAD 图纸单位按 mm 处理，闭合多段线面积换算为 m2；可用 `--unit-scale` 调整。",
            "- 单位冷负荷：按房间名称经验分类生成建议，正式负荷需设计师复核。",
            "",
            "## 识别结果",
            "",
            f"- 房间候选：{len(rooms)} 个。",
            f"- 已形成负荷计算：{len(calculated)} 个。",
            f"- 已计算总冷负荷：{total_load:.2f} kW。",
            "",
            "## 室内机候选",
            "",
        ]
    )
    for row in selections:
        capacity = "" if row.selected_capacity_kw is None else f"{row.quantity}x{row.selected_capacity_kw}kW"
        lines.append(f"- {row.room_id} {row.room_name}: {capacity or '待面积确认'}，{row.indoor_unit_form}，{row.conclusion}")
    lines.extend(["", "## 待确认事项", ""])
    if issues:
        lines.extend(f"- {row.issue_id} [{row.severity}] {row.description}：{row.recommendation}" for row in issues)
    else:
        lines.append("- 暂无自动发现问题；仍需设计师复核图纸边界、负荷口径和厂家选型规则。")
    path.write_text("\n".join(lines), encoding="utf-8-sig")


def source_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".dwg", ".dxf", ".dwf", ".dwfx", ".pdf"}:
        return "drawing"
    if suffix in {".xlsx", ".xls", ".xlsm", ".csv"}:
        return "table"
    return "unknown"


def register_sources(inputs: list[Path]) -> list[SourceRow]:
    rows: list[SourceRow] = []
    for index, path in enumerate(inputs, start=1):
        status = "used" if path.suffix.lower() == ".dxf" else "pending conversion"
        if not path.exists():
            status = "missing"
        rows.append(
            SourceRow(
                source_id=f"SRC-{index:03d}",
                file_or_link=str(path),
                type=source_type(path),
                version_date=datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds") if path.exists() else "",
                owner_role="automation-engineer",
                status=status,
                notes="DXF 可直接读取；DWG/DWF/DWFX/PDF 需转换或 OCR 后进入房间识别",
                sha256=sha256(path) if path.exists() else "",
            )
        )
    return rows


def process(args: argparse.Namespace) -> None:
    inputs = [Path(item).resolve() for item in args.input]
    out_dir = (Path(args.output_dir).resolve() if args.output_dir else DEFAULT_OUTPUT_ROOT / args.task_name)
    out_dir.mkdir(parents=True, exist_ok=True)

    source_rows = register_sources(inputs)
    all_texts: list[TextRow] = []
    all_rooms: list[RoomRow] = []

    for row, path in zip(source_rows, inputs):
        if row.status == "missing" or path.suffix.lower() != ".dxf":
            continue
        doc = read_dxf(path)
        texts = extract_texts(doc, row.source_id)
        polygons = extract_room_polygons(
            doc,
            min_area=args.min_room_area_m2 / (args.unit_scale * args.unit_scale),
            max_area=args.max_room_area_m2 / (args.unit_scale * args.unit_scale),
        )
        rooms = build_rooms(row.source_id, texts, polygons, args.unit_scale, args.correction_factor)
        all_texts.extend(texts)
        all_rooms.extend(rooms)

    selections = build_selections(all_rooms, args.brand, args.series)
    issues = build_issues(inputs, all_rooms, args.brand, args.city)

    write_csv(out_dir / "source_register.csv", source_rows)
    write_csv(out_dir / "text_entities.csv", all_texts)
    write_csv(out_dir / "room_load_draft.csv", all_rooms)
    write_csv(out_dir / "indoor_unit_selection_draft.csv", selections)
    write_csv(out_dir / "issue_ledger.csv", issues)
    write_workbook(
        out_dir / "vrf_design_quote_draft.xlsx",
        {
            "资料登记": source_rows,
            "文字图元": all_texts,
            "房间负荷底稿": all_rooms,
            "室内机选型候选": selections,
            "待确认事项": issues,
        },
    )
    write_overlay_dxf(out_dir / "vrf_indoor_unit_overlay_draft.dxf", all_rooms, selections)
    write_report(
        out_dir / "vrf_design_quote_report.md",
        source_rows,
        all_rooms,
        selections,
        issues,
        args.brand,
        args.series,
        args.city,
    )
    print(f"output_dir={out_dir}")
    print(f"sources={len(source_rows)}")
    print(f"texts={len(all_texts)}")
    print(f"rooms={len(all_rooms)}")
    print(f"selections={len(selections)}")
    print(f"issues={len(issues)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract VRF room/load/selection draft from explicitly provided CAD inputs.")
    parser.add_argument("--input", action="append", required=True, help="Explicit input drawing path. Repeat for multiple files.")
    parser.add_argument("--task-name", default="vrf_design_pipeline", help="Output subdirectory under outputs/.")
    parser.add_argument("--output-dir", default="", help="Optional explicit output directory.")
    parser.add_argument("--city", default="", help="Project city for later weather/design-parameter traceability.")
    parser.add_argument("--brand", default="通用", help="Candidate brand. Use manufacturer manuals before finalizing.")
    parser.add_argument("--series", default="通用", help="Candidate series.")
    parser.add_argument("--unit-scale", type=float, default=0.001, help="CAD drawing unit to meter scale. Default 0.001 for mm.")
    parser.add_argument("--correction-factor", type=float, default=1.0, help="Load correction factor applied to all rooms.")
    parser.add_argument("--min-room-area-m2", type=float, default=3.0, help="Minimum closed polyline area treated as room.")
    parser.add_argument("--max-room-area-m2", type=float, default=500.0, help="Maximum closed polyline area treated as room.")
    return parser.parse_args()


def main() -> None:
    process(parse_args())


if __name__ == "__main__":
    main()
