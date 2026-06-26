from __future__ import annotations

import csv
import math
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

import ezdxf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DXF_PATH = ROOT / "outputs" / "dwg_conversion_test" / "7F_decorated_ceiling_plan_R2018.dxf"
OUT_DIR = ROOT / "outputs" / "xitou_yuncheng_7f_hvac_cross_conditions"

SOURCE_DWG = (
    "C:\\Users\\Liugq\\Desktop\\施工图-西投云城犀谷7#楼7-8层装修工程(1)\\"
    "施工图-西投云城犀谷7#楼7-8层装修工程\\"
    "西投云城犀谷7号楼7F装饰+电气施工图（2026.04）\\7F装饰\\"
    "01西投云城犀谷7#楼7F平顶面图\\"
    "01-西投云城犀谷7#楼7F装饰平顶面施工图(2026.04).dwg"
)

HVAC_KEYWORDS = ["新风", "冷媒", "送风", "回风", "空调", "管井", "冷媒井", "风口", "排风"]
CEILING_KEYWORDS = ["吊顶", "天花", "平顶", "检修", "留洞", "标高", "贴梁底", "灯具"]
INDOOR_UNIT_LAYERS = {"暖通-多联机-室内机"}
CEILING_DEVICE_LAYERS = {"IRC-WIR　顶面设备"}


def clean_cad_text(value: str) -> str:
    text = str(value or "")
    text = text.replace("\\P", " ").replace("\n", " ")
    text = re.sub(r"\{\\[^;{}]+;", "", text)
    text = text.replace("{", "").replace("}", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


@dataclass
class TextEntity:
    entity_type: str
    layer: str
    text: str
    x: float
    y: float
    source: str


@dataclass
class HvacPoint:
    point_id: str
    category: str
    keyword: str
    text: str
    layer: str
    x: float
    y: float
    nearby_text: str
    evidence: str
    review_status: str


@dataclass
class LayerFinding:
    layer: str
    reason: str
    entity_count: int
    text_count: int
    sample_text: str
    recommended_use: str


@dataclass
class NearbyInsert:
    point_id: str
    insert_name: str
    insert_layer: str
    distance: float
    x: float
    y: float
    evidence: str


@dataclass
class LayoutCandidate:
    candidate_id: str
    candidate_type: str
    source_priority: str
    layer: str
    insert_name: str
    x: float
    y: float
    rotation: float
    xscale: float
    yscale: float
    evidence: str
    review_status: str


@dataclass
class LayoutContext:
    candidate_id: str
    candidate_type: str
    insert_name: str
    x: float
    y: float
    nearest_layout_candidates: str
    nearby_room_or_function_text: str
    evidence: str
    review_status: str


@dataclass
class Issue:
    issue_id: str
    severity: str
    source: str
    description: str
    evidence: str
    recommendation: str
    status: str


def text_of(entity) -> str:
    if entity.dxftype() == "TEXT":
        return clean_cad_text(entity.dxf.text)
    if entity.dxftype() == "MTEXT":
        if hasattr(entity, "plain_text"):
            return clean_cad_text(entity.plain_text())
        return clean_cad_text(entity.text)
    return ""


def location_of(entity) -> tuple[float, float]:
    loc = getattr(entity.dxf, "insert", None)
    if loc is None:
        return 0.0, 0.0
    return float(loc.x), float(loc.y)


def iter_texts(msp) -> Iterable[TextEntity]:
    for entity in msp:
        if entity.dxftype() not in {"TEXT", "MTEXT"}:
            continue
        text = text_of(entity)
        if not text:
            continue
        x, y = location_of(entity)
        yield TextEntity(
            entity_type=entity.dxftype(),
            layer=entity.dxf.layer,
            text=text,
            x=x,
            y=y,
            source=f"{entity.dxftype()}@{entity.dxf.layer}({x:.3f},{y:.3f})",
        )


def category_for(text: str) -> tuple[str, str] | None:
    for keyword in HVAC_KEYWORDS:
        if keyword in text:
            if "冷媒" in keyword or "冷媒" in text:
                return "冷媒/冷媒井候选", keyword
            if "新风" in keyword or "新风" in text:
                return "新风候选", keyword
            if "送风" in keyword or "送风" in text:
                return "送风候选", keyword
            if "回风" in keyword or "回风" in text:
                return "回风候选", keyword
            if "管井" in keyword or "管井" in text:
                return "管井候选", keyword
            return "暖通相关候选", keyword
    for keyword in CEILING_KEYWORDS:
        if keyword in text:
            return "吊顶/检修条件候选", keyword
    return None


def dist(a: TextEntity, b: TextEntity) -> float:
    return math.hypot(a.x - b.x, a.y - b.y)


def nearby_context(target: TextEntity, texts: list[TextEntity], radius: float = 4500.0) -> str:
    nearby = []
    for other in texts:
        if other is target:
            continue
        d = dist(target, other)
        if d <= radius:
            nearby.append((d, other))
    nearby.sort(key=lambda item: item[0])
    return " | ".join(f"{item.text}@{item.layer},d={d:.0f}" for d, item in nearby[:8])


def nearby_text_context(x: float, y: float, texts: list[TextEntity], radius: float = 5500.0) -> str:
    nearby = []
    for text in texts:
        d = math.hypot(x - text.x, y - text.y)
        if d <= radius:
            nearby.append((d, text))
    nearby.sort(key=lambda item: item[0])
    return " | ".join(f"{item.text}@{item.layer},d={d:.0f}" for d, item in nearby[:10])


def collect_inserts(msp) -> list[dict[str, object]]:
    inserts = []
    for entity in msp:
        if entity.dxftype() != "INSERT":
            continue
        loc = entity.dxf.insert
        inserts.append(
            {
                "name": entity.dxf.name,
                "layer": entity.dxf.layer,
                "x": float(loc.x),
                "y": float(loc.y),
            }
        )
    return inserts


def collect_layout_candidates(msp) -> list[LayoutCandidate]:
    candidates: list[LayoutCandidate] = []
    for entity in msp:
        if entity.dxftype() != "INSERT":
            continue
        layer = entity.dxf.layer
        if layer in INDOOR_UNIT_LAYERS:
            candidate_type = "室内机布置候选"
            source_priority = "优先按装饰/暖通协调布局复核"
        elif layer in CEILING_DEVICE_LAYERS:
            candidate_type = "顶面设备/风口候选"
            source_priority = "优先作为出回风口或顶面设备布置输入，需人工确认符号含义"
        else:
            continue
        loc = entity.dxf.insert
        candidate_id = f"LAYOUT-{len(candidates) + 1:03d}"
        candidates.append(
            LayoutCandidate(
                candidate_id=candidate_id,
                candidate_type=candidate_type,
                source_priority=source_priority,
                layer=layer,
                insert_name=entity.dxf.name,
                x=round(float(loc.x), 3),
                y=round(float(loc.y), 3),
                rotation=round(float(getattr(entity.dxf, "rotation", 0.0) or 0.0), 3),
                xscale=round(float(getattr(entity.dxf, "xscale", 1.0) or 1.0), 3),
                yscale=round(float(getattr(entity.dxf, "yscale", 1.0) or 1.0), 3),
                evidence=f"INSERT:{entity.dxf.name}@{layer}({float(loc.x):.3f},{float(loc.y):.3f})",
                review_status="pending symbol and room mapping review",
            )
        )
    return candidates


def collect_layout_contexts(candidates: list[LayoutCandidate], texts: list[TextEntity]) -> list[LayoutContext]:
    contexts: list[LayoutContext] = []
    for candidate in candidates:
        neighbors = []
        for other in candidates:
            if other.candidate_id == candidate.candidate_id:
                continue
            d = math.hypot(candidate.x - other.x, candidate.y - other.y)
            if d <= 5500:
                neighbors.append((d, other))
        neighbors.sort(key=lambda item: item[0])
        nearest = " | ".join(
            f"{other.candidate_id}:{other.candidate_type}:{other.insert_name}@{other.layer},d={d:.0f}"
            for d, other in neighbors[:8]
        )
        contexts.append(
            LayoutContext(
                candidate_id=candidate.candidate_id,
                candidate_type=candidate.candidate_type,
                insert_name=candidate.insert_name,
                x=candidate.x,
                y=candidate.y,
                nearest_layout_candidates=nearest,
                nearby_room_or_function_text=nearby_text_context(candidate.x, candidate.y, texts),
                evidence=f"{candidate.evidence}; neighbor radius=5500 drawing units",
                review_status="pending room boundary and symbol review",
            )
        )
    return contexts


def nearest_inserts(point: HvacPoint, inserts: list[dict[str, object]], limit: int = 5, radius: float = 2500.0) -> list[NearbyInsert]:
    candidates = []
    for insert in inserts:
        d = math.hypot(point.x - float(insert["x"]), point.y - float(insert["y"]))
        if d <= radius:
            candidates.append((d, insert))
    candidates.sort(key=lambda item: item[0])
    return [
        NearbyInsert(
            point_id=point.point_id,
            insert_name=str(insert["name"]),
            insert_layer=str(insert["layer"]),
            distance=round(d, 3),
            x=float(insert["x"]),
            y=float(insert["y"]),
            evidence=f"{point.point_id} near INSERT:{insert['name']}@{insert['layer']}",
        )
        for d, insert in candidates[:limit]
    ]


def write_csv(path: Path, rows: list[object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    fields = list(asdict(rows[0]).keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def add_sheet(wb: Workbook, title: str, rows: list[object]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["无数据"])
        return
    fields = list(asdict(rows[0]).keys())
    ws.append(fields)
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    for row in rows:
        ws.append([asdict(row)[field] for field in fields])
    ws.freeze_panes = "A2"
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 70)
        ws.column_dimensions[column[0].column_letter].width = width


def main() -> None:
    if not DXF_PATH.exists():
        raise FileNotFoundError(f"DXF not found: {DXF_PATH}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    doc = ezdxf.readfile(DXF_PATH)
    msp = doc.modelspace()
    texts = list(iter_texts(msp))
    inserts = collect_inserts(msp)
    layout_candidates = collect_layout_candidates(msp)
    layout_contexts = collect_layout_contexts(layout_candidates, texts)

    points: list[HvacPoint] = []
    for text in texts:
        category_keyword = category_for(text.text)
        if category_keyword is None:
            continue
        category, keyword = category_keyword
        point_id = f"HVAC-{len(points) + 1:03d}"
        points.append(
            HvacPoint(
                point_id=point_id,
                category=category,
                keyword=keyword,
                text=text.text,
                layer=text.layer,
                x=round(text.x, 3),
                y=round(text.y, 3),
                nearby_text=nearby_context(text, texts),
                evidence=text.source,
                review_status="pending manual review",
            )
        )

    layer_entity_counts: dict[str, int] = {}
    for entity in msp:
        layer = getattr(entity.dxf, "layer", "")
        if layer:
            layer_entity_counts[layer] = layer_entity_counts.get(layer, 0) + 1

    layer_texts: dict[str, list[str]] = {}
    for text in texts:
        layer_texts.setdefault(text.layer, []).append(text.text)

    layer_findings: list[LayerFinding] = []
    for layer, samples in sorted(layer_texts.items()):
        joined = " ".join(samples)
        matched = [kw for kw in HVAC_KEYWORDS + CEILING_KEYWORDS if kw in joined or kw in layer]
        if not matched:
            continue
        reason = ", ".join(sorted(set(matched)))
        recommended = "暖通交叉条件人工复核" if any(kw in reason for kw in HVAC_KEYWORDS) else "吊顶/检修条件人工复核"
        layer_findings.append(
            LayerFinding(
                layer=layer,
                reason=reason,
                entity_count=layer_entity_counts.get(layer, 0),
                text_count=len(samples),
                sample_text=" | ".join(samples[:10]),
                recommended_use=recommended,
            )
        )

    nearby_inserts: list[NearbyInsert] = []
    for point in points:
        if "吊顶" in point.category or "检修" in point.category:
            continue
        nearby_inserts.extend(nearest_inserts(point, inserts))

    issues = [
        Issue(
            issue_id="QA-001",
            severity="S2",
            source="design",
            description="本图为装饰平顶面图，不是暖通主图，不能直接作为多联机设备选型、系统划分或管径设计依据。",
            evidence=f"source DWG: {SOURCE_DWG}",
            recommendation="补充暖通/空调施工图、设备表、系统图和项目设计参数；本底稿仅用于交叉条件识别。",
            status="open",
        ),
        Issue(
            issue_id="QA-002",
            severity="S3",
            source="automation",
            description="已识别新风、冷媒、送风、冷媒井等文字候选，但未解析其对应管井边界或房间归属。",
            evidence="hvac_cross_condition_points.csv",
            recommendation="下一步结合墙体/房间边界图层和人工复核，建立管井候选与房间/区域的对应关系。",
            status="open",
        ),
        Issue(
            issue_id="QA-003",
            severity="S3",
            source="design",
            description="已单独抽取装饰/暖通协调图层中的室内机和顶面设备候选；按当前规则应优先作为室内机、出风口、回风口布局输入。",
            evidence="layout_candidates.csv",
            recommendation="下一步应确认顶面设备块符号含义，并校核是否满足暖通规范、使用效果和售后维修维护最低要求。",
            status="open",
        ),
    ]

    write_csv(OUT_DIR / "hvac_cross_condition_points.csv", points)
    write_csv(OUT_DIR / "layout_candidates.csv", layout_candidates)
    write_csv(OUT_DIR / "layout_contexts.csv", layout_contexts)
    write_csv(OUT_DIR / "hvac_related_layers.csv", layer_findings)
    write_csv(OUT_DIR / "nearby_insert_candidates.csv", nearby_inserts)
    write_csv(OUT_DIR / "issue_ledger.csv", issues)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "暖通交叉点位", points)
    add_sheet(wb, "装饰布局候选", layout_candidates)
    add_sheet(wb, "布局邻近关系", layout_contexts)
    add_sheet(wb, "相关图层", layer_findings)
    add_sheet(wb, "附近图块候选", nearby_inserts)
    add_sheet(wb, "问题台账", issues)
    wb.save(OUT_DIR / "7f_hvac_cross_conditions.xlsx")

    report = [
        "# 西投云城 7F 平顶面暖通交叉条件底稿",
        "",
        f"- 来源 DWG：`{SOURCE_DWG}`",
        f"- 派生 DXF：`{DXF_PATH.relative_to(ROOT)}`",
        "- 原始 DWG 未修改；本底稿基于派生 DXF 只读抽取。",
        "",
        "## 抽取结果",
        "",
        f"- 暖通/吊顶相关文字点位：{len(points)} 条",
        f"- 装饰布局候选：{len(layout_candidates)} 条",
        f"- 布局邻近关系：{len(layout_contexts)} 条",
        f"- 相关图层：{len(layer_findings)} 个",
        f"- 附近图块候选：{len(nearby_inserts)} 条",
        "",
        "## 关键观察",
        "",
        "- 图中存在 `新风`、`冷媒`、`送风`、`冷媒井` 等文字，可作为暖通交叉条件候选。",
        "- 图中存在 `暖通-多联机-室内机` 和 `IRC-WIR　顶面设备` 图层的块参照，已按“装饰出/回风口和室内机布局优先”规则单独登记。",
        "- 这些候选主要来自装饰平顶面图中的管井文字/相关文字，不等同于暖通主图设计结论。",
        "- 需要结合暖通施工图、设备表、系统图、房间边界、块符号说明和人工复核，才能继续做选型、系统划分和管路设计。",
        "",
        "## 输出文件",
        "",
        "- `hvac_cross_condition_points.csv`：暖通/吊顶相关文字点位、坐标、附近文字上下文",
        "- `layout_candidates.csv`：室内机、顶面设备/风口候选块参照",
        "- `layout_contexts.csv`：室内机/顶面设备候选之间的邻近关系和附近房间/功能文字",
        "- `hvac_related_layers.csv`：相关图层和样本文字",
        "- `nearby_insert_candidates.csv`：点位附近图块候选",
        "- `7f_hvac_cross_conditions.xlsx`：汇总工作簿",
        "- `issue_ledger.csv`：问题台账",
    ]
    (OUT_DIR / "7f_hvac_cross_conditions_report.md").write_text("\n".join(report), encoding="utf-8-sig")

    print(f"points={len(points)} layouts={len(layout_candidates)} layers={len(layer_findings)} inserts={len(nearby_inserts)}")
    print(f"output={OUT_DIR}")


if __name__ == "__main__":
    main()
