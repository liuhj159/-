from __future__ import annotations

import csv
import hashlib
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "outputs" / "feishu_quote_template_import"
IMPORT_XLSX = OUT_DIR / "5_brand_quote_template_feishu_import.xlsx"


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def find_source_template() -> Path:
    desktop = Path.home() / "Desktop"
    candidates = sorted(desktop.glob("5*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError("桌面未找到以 5 开头的报价模板 xlsx 文件。")
    return candidates[0]


def inspect_workbook(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, read_only=False, data_only=False)
    rows: list[dict[str, object]] = []
    for ws in wb.worksheets:
        nonempty = 0
        formulas = 0
        headers = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                nonempty += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas += 1
        for cell in ws[1]:
            if cell.value is not None:
                headers.append(str(cell.value))
        if "报价" in ws.title:
            suggested_use = "报价填写/输出"
        elif "价格数据" in ws.title:
            suggested_use = "品牌价格数据源"
        else:
            suggested_use = "模板导航/辅助表"
        rows.append(
            {
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "nonempty_cells": nonempty,
                "formula_cells": formulas,
                "suggested_feishu_use": suggested_use,
                "header_sample": " | ".join(headers[:12]),
            }
        )
    return rows


def write_manifest(rows: list[dict[str, object]]) -> None:
    path = OUT_DIR / "sheet_manifest.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_report(source: Path, rows: list[dict[str, object]]) -> None:
    report = [
        "# 飞书报价模板导入包",
        "",
        "## 源文件",
        "",
        f"- 桌面源模板：`{source}`",
        f"- 飞书导入源：`{IMPORT_XLSX}`",
        f"- 源文件 SHA256：`{sha256(source)}`",
        f"- 导入源 SHA256：`{sha256(IMPORT_XLSX)}`",
        "",
        "## 建议飞书结构",
        "",
        "优先导入为飞书电子表格，保持原 Excel 的多工作表和公式结构；暂不建议直接拆成多维表格，因为当前模板包含大量跨表公式和品牌价格数据表。",
        "",
        "导入后建议命名：`5品牌多联机报价模板`。",
        "",
        "## 工作表清单",
        "",
    ]
    for row in rows:
        report.append(
            f"- `{row['sheet_name']}`：{row['max_row']} 行 x {row['max_column']} 列，"
            f"公式 {row['formula_cells']} 个，用途：{row['suggested_feishu_use']}"
        )
    report.extend(
        [
            "",
            "## 后续图纸方案上传流程",
            "",
            "1. 图纸自动化流程输出设备表、工程量表、系统配置表和待确认事项。",
            "2. 将输出表上传到飞书项目资料夹。",
            "3. 复制本报价模板生成项目报价副本。",
            "4. 把设备型号、数量、工程量、辅材和人工口径写入对应品牌报价页。",
            "5. 报价师复核库存、单价、税费、利润和替代型号。",
            "6. 审计复核公式、单位、来源和版本后定稿。",
            "",
            "## 安全边界",
            "",
            "- 未配置 `.env` 或 `ALLOW_EXTERNAL_UPLOAD=true` 前，不自动上传到飞书。",
            "- 真实密钥不得写入本报告、脚本、Excel、CSV 或 Git 提交。",
        ]
    )
    (OUT_DIR / "feishu_upload_plan.md").write_text("\n".join(report), encoding="utf-8-sig")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    source = find_source_template()
    copy2(source, IMPORT_XLSX)
    rows = inspect_workbook(source)
    write_manifest(rows)
    write_report(source, rows)
    print(f"source={source}")
    print(f"import_xlsx={IMPORT_XLSX}")
    print(f"sheets={len(rows)}")
    print(f"output={OUT_DIR}")


if __name__ == "__main__":
    main()
