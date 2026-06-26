from __future__ import annotations

import csv
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
INPUT_CSV = ROOT / "outputs" / "6-12_plan_scan" / "room_load_draft.csv"
OUT_DIR = ROOT / "outputs" / "6-12_area_completion"


# area_m2, perimeter_m, method, evidence
AREA_BY_INDEX: dict[int, tuple[float, float | str, str, str]] = {
    1: (26.35, 20.16, "estimated_from_cad_wall_axis", "x:-334106~-327296;y:1051816~1055686"),
    2: (124.17, 44.64, "estimated_from_cad_wall_axis", "x:-308756~-298206;y:1052516~1064286"),
    3: (36.59, 25.28, "estimated_from_cad_wall_axis", "x:-334106~-325956;y:1055686~1060176"),
    4: (37.02, 25.74, "estimated_from_cad_wall_axis", "x:-334106~-325576;y:1060176~1064516"),
    5: (33.74, 24.97, "estimated_from_cad_wall_axis", "x:-334106~-325576;y:1064516~1068471"),
    6: (45.17, 31.97, "confirmed_by_user_autocad_area", "user screenshot: AREA=45168106.11; perimeter=31970.00"),
    7: (13.09, 15.53, "estimated_from_cad_wall_axis", "x:-303496~-298206;y:1071391~1073866"),
    8: (37.61, 24.83, "estimated_from_cad_wall_axis", "x:-303496~-298206;y:1064286~1071391"),
    9: (20.39, 18.15, "estimated_from_cad_wall_axis", "x:-324406~-320311;y:1068886~1073866"),
    10: (20.42, 18.16, "estimated_from_cad_wall_axis", "x:-320311~-316211;y:1068886~1073866"),
    11: (23.64, 19.48, "estimated_from_cad_wall_axis", "x:-308636~-303496;y:1064286~1068886"),
    12: (21.56, 18.62, "estimated_from_cad_wall_axis", "x:-316211~-311881;y:1068886~1073866"),
    13: (19.62, 17.84, "estimated_from_cad_wall_axis", "x:-321956~-318016;y:1061866~1066846"),
    14: (4.56, 8.90, "estimated_from_cad_wall_axis", "x:-299806~-298206;y:1052716~1055566"),
    15: (12.88, 19.30, "estimated_from_cad_wall_axis", "x:-324406~-316356;y:1029566~1031166"),
    16: (23.28, 19.31, "estimated_from_cad_wall_axis", "x:-311881~-307206;y:1068886~1073866"),
    17: (11.27, 13.44, "estimated_from_cad_wall_axis", "x:-303631~-300406;y:1034991~1038486"),
    18: (37.42, 24.97, "estimated_from_cad_wall_axis", "x:-334106~-326616;y:1038366~1043361"),
    19: (31.53, 23.40, "estimated_from_cad_wall_axis", "x:-334106~-326616;y:1043361~1047571"),
    20: (5.96, 10.41, "estimated_from_cad_wall_axis", "x:-323656~-320156;y:1060163~1061866"),
    21: (8.67, 13.60, "estimated_from_cad_wall_axis", "x:-323656~-321956;y:1061866~1066966"),
    22: (4.10, "", "furniture_cluster_bbox_pending_room_confirmation", "unnamed_furniture_room_candidates.csv bbox; not net room area"),
}

CAPACITIES_KW = [2.2, 2.8, 3.6, 4.5, 5.6, 7.1, 8.0, 9.0, 11.2, 14.0, 16.0, 22.4, 28.0]


def select_capacity(load_kw: float) -> tuple[float, int, float]:
    for capacity in CAPACITIES_KW:
        if capacity >= load_kw:
            return capacity, 1, (capacity / load_kw - 1) * 100
    qty = math.ceil(load_kw / 14.0)
    return 14.0, qty, (qty * 14.0 / load_kw - 1) * 100


def unit_form(function: str, load_kw: float) -> str:
    if any(key in function for key in ["会议", "包厢", "会客", "接待"]):
        return "四面出风嵌入式或薄型风管机候选"
    if "清洗" in function:
        return "薄型风管机候选，需复核排水和吊顶条件"
    if load_kw >= 12:
        return "多台薄型风管机或高静压风管机候选"
    return "薄型风管机候选"


def build_rows() -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    with INPUT_CSV.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for source in reader:
            idx = int(source["序号"])
            area, perimeter, method, evidence = AREA_BY_INDEX[idx]
            unit_load = float(source["单位冷负荷W/m2建议"])
            load_w = area * unit_load
            load_kw = load_w / 1000
            capacity, qty, margin = select_capacity(load_kw)
            if method == "furniture_cluster_bbox_pending_room_confirmation":
                status = "家具聚类候选，待确认是否为房间"
            elif method.startswith("confirmed"):
                status = "已按用户AREA示范确认"
            else:
                status = "CAD坐标估算，待人工AREA复核"
            rows.append(
                {
                    "序号": idx,
                    "房间名称": source["房间名称"],
                    "房间功能建议": source["房间功能建议"],
                    "面积m2": round(area, 2),
                    "周长m": perimeter,
                    "单位冷负荷W/m2建议": unit_load,
                    "修正系数": 1.0,
                    "房间冷负荷W": round(load_w, 1),
                    "房间冷负荷kW": round(load_kw, 2),
                    "候选室内机容量kW": capacity,
                    "数量": qty,
                    "容量余量%": round(margin, 1),
                    "室内机形式候选": unit_form(source["房间功能建议"], load_kw),
                    "面积来源方法": method,
                    "证据/坐标依据": evidence,
                    "状态": status,
                    "备注": "估算值可被后续 AutoCAD AREA 人工测量结果覆盖",
                }
            )
    return rows


def write_csv(rows: list[dict[str, object]]) -> Path:
    path = OUT_DIR / "room_area_load_selection_completed.csv"
    fields = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    return path


def write_xlsx(rows: list[dict[str, object]]) -> Path:
    path = OUT_DIR / "6-12房间面积负荷选型补全.xlsx"
    fields = list(rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "面积负荷选型补全"
    ws.append(fields)
    for row in rows:
        ws.append([row[field] for field in fields])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, field in enumerate(fields, start=1):
        ws.column_dimensions[get_column_letter(index)].width = min(max(len(field) + 4, 12), 36)
    ws.freeze_panes = "A2"

    total_area = sum(float(row["面积m2"]) for row in rows if row["面积来源方法"] != "furniture_cluster_bbox_pending_room_confirmation")
    total_load = sum(float(row["房间冷负荷kW"]) for row in rows if row["面积来源方法"] != "furniture_cluster_bbox_pending_room_confirmation")
    summary = wb.create_sheet("汇总")
    for row in [
        ["项目", "值"],
        ["补全面积房间数", len(rows)],
        ["扣除未命名家具候选后的面积合计m2", round(total_area, 2)],
        ["扣除未命名家具候选后的冷负荷合计kW", round(total_load, 2)],
        ["已确认面积", "招标部 45.17m2"],
        ["其余状态", "CAD坐标估算，待人工AREA复核"],
    ]:
        summary.append(row)
    for cell in summary[1]:
        cell.font = Font(bold=True)
    wb.save(path)
    return path


def write_report(rows: list[dict[str, object]]) -> Path:
    total_area = sum(float(row["面积m2"]) for row in rows if row["面积来源方法"] != "furniture_cluster_bbox_pending_room_confirmation")
    total_load = sum(float(row["房间冷负荷kW"]) for row in rows if row["面积来源方法"] != "furniture_cluster_bbox_pending_room_confirmation")
    path = OUT_DIR / "6-12_area_completion_report.md"
    lines = [
        "# 6-12平面房间面积补全底稿",
        "",
        "## 口径",
        "",
        "- 原始图纸保持只读。",
        "- 招标部采用用户 AutoCAD AREA 示范结果：面积 45168106.11、周长 31970.00，按 mm 图换算为 45.17m2、31.97m。",
        "- 其余房间按 DXF 墙线/外墙线坐标推算服务区域面积，状态均为待人工 AREA 复核。",
        "- 家具、柜体、桌椅不作为边界；门洞按墙体边界补齐。",
        "",
        "## 汇总",
        "",
        f"- 补全记录：{len(rows)} 条。",
        f"- 扣除未命名家具候选后的面积合计：{total_area:.2f} m2。",
        f"- 扣除未命名家具候选后的冷负荷合计：{total_load:.2f} kW。",
        "",
        "## 待复核",
        "",
        "- 除招标部外，其余面积为 CAD 坐标估算值，应按用户示范的 AutoCAD AREA 点选法逐房间复核。",
        "- R-6-12-WM01 来自家具聚类包围盒，不作为正式房间面积，只保留为待确认服务区域。",
        "",
        "## 输出文件",
        "",
        "- room_area_load_selection_completed.csv",
        "- 6-12房间面积负荷选型补全.xlsx",
    ]
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    return path


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = build_rows()
    csv_path = write_csv(rows)
    xlsx_path = write_xlsx(rows)
    report_path = write_report(rows)
    print(f"rows={len(rows)}")
    print(f"csv={csv_path}")
    print(f"xlsx={xlsx_path}")
    print(f"report={report_path}")


if __name__ == "__main__":
    main()
