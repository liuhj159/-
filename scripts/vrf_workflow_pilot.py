from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import mimetypes
import re
import shutil
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT_DIR = ROOT / "outputs" / "auto_workflow_pilot"
FIXTURE_ROOT = ROOT / "tests" / "fixtures"


@dataclass
class SourceRow:
    source_id: str
    file_or_link: str
    type: str
    version_date: str
    owner_role: str
    status: str
    notes: str
    sha256: str


@dataclass
class AutomationRow:
    source_id: str
    file_or_link: str
    format: str
    readable_now: str
    extraction_route: str
    extracted_artifacts: str
    confidence: str
    manual_review_point: str
    next_tool_need: str


@dataclass
class IssueRow:
    issue_id: str
    severity: str
    source: str
    description: str
    impact: str
    evidence: str
    owner: str
    recommendation: str
    status: str
    release_conclusion: str


@dataclass
class WorkbookSheetRow:
    source_id: str
    file_or_link: str
    sheet_name: str
    max_row: int
    max_column: int


@dataclass
class MaterialRow:
    source_id: str
    file_or_link: str
    sheet_name: str
    row_number: int
    item_code: str
    category: str
    item_name: str
    dimension_or_param: str
    material: str
    specification: str
    installation_location: str
    notes: str


@dataclass
class DwgIndexRow:
    source_id: str
    file_or_link: str
    dwg_header: str
    acad_version: str
    file_size_bytes: int
    can_parse_geometry_now: str
    next_step: str


@dataclass
class DrawingDisciplineRow:
    source_id: str
    file_or_link: str
    discipline: str
    evidence_keyword: str
    hvac_relevant: str
    recommended_use: str


@dataclass
class MaterialCompareRow:
    sheet_name: str
    item_code: str
    source_ids: str
    item_names: str
    dimensions_or_params: str
    installation_locations: str
    conclusion: str
    evidence: str


@dataclass
class VrfSystemConfigRow:
    source_id: str
    file_or_link: str
    sheet_name: str
    row_number: int
    floor_or_area: str
    room_name: str
    area_m2: str
    unit_load_w_m2: str
    equipment_type: str
    equipment_model: str
    capacity_kw: str
    quantity: str
    cooling_capacity_kw: str
    connection_ratio: str
    system_mark: str
    evidence: str


@dataclass
class VrfEquipmentSummaryRow:
    source_id: str
    file_or_link: str
    sheet_name: str
    row_number: int
    equipment_name: str
    equipment_model: str
    quantity: str
    unit: str
    unit_price: str
    total_price: str
    evidence: str


@dataclass
class VrfInstallMaterialRow:
    source_id: str
    file_or_link: str
    sheet_name: str
    row_number: int
    item_name: str
    spec: str
    brand: str
    quantity: str
    unit: str
    unit_price: str
    total_price: str
    notes: str
    floor_breakdown: str
    evidence: str


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


DWG_VERSION_MAP = {
    "AC1009": "AutoCAD R12",
    "AC1012": "AutoCAD R13",
    "AC1014": "AutoCAD R14",
    "AC1015": "AutoCAD 2000/2000i/2002",
    "AC1018": "AutoCAD 2004/2005/2006",
    "AC1021": "AutoCAD 2007/2008/2009",
    "AC1024": "AutoCAD 2010/2011/2012",
    "AC1027": "AutoCAD 2013/2014/2015/2016/2017",
    "AC1032": "AutoCAD 2018/2019/2020/2021/2022/2023/2024/2025/2026",
}


def inspect_dwg_header(source_id: str, path: Path) -> DwgIndexRow:
    with path.open("rb") as f:
        header = f.read(6).decode("ascii", errors="replace")
    return DwgIndexRow(
        source_id=source_id,
        file_or_link=relative_label(path),
        dwg_header=header,
        acad_version=DWG_VERSION_MAP.get(header, "unknown"),
        file_size_bytes=path.stat().st_size,
        can_parse_geometry_now="no" if not oda_installed() else "yes",
        next_step="安装/验证 ODA File Converter 后转 DXF，再用 ezdxf 抽取图元" if not oda_installed() else "运行 DWG -> DXF -> ezdxf 图元抽取",
    )


def classify_drawing_discipline(source_id: str, path: Path) -> DrawingDisciplineRow:
    text = f"{path.parent.name} {path.name}"
    checks = [
        ("暖通/空调", ["暖通", "空调", "多联机", "通风", "防排烟"], "yes", "可作为暖通/多联机设计校核输入"),
        ("电气", ["电气", "强电", "弱电", "照明"], "partial", "可用于电源、控制线和灯具相关交叉核对"),
        ("给排水", ["给排水", "排水", "给水"], "partial", "可用于冷凝水排水点和专业交叉核对"),
        ("装饰", ["装饰", "室内", "平顶面", "立面", "节点", "门表", "材料"], "partial", "可用于吊顶、检修口、室内点位和材料交叉核对"),
    ]
    for discipline, keywords, hvac_relevant, recommended_use in checks:
        for keyword in keywords:
            if keyword in text:
                return DrawingDisciplineRow(
                    source_id=source_id,
                    file_or_link=relative_label(path),
                    discipline=discipline,
                    evidence_keyword=keyword,
                    hvac_relevant=hvac_relevant,
                    recommended_use=recommended_use,
                )
    return DrawingDisciplineRow(
        source_id=source_id,
        file_or_link=relative_label(path),
        discipline="unknown",
        evidence_keyword="",
        hvac_relevant="pending",
        recommended_use="需人工确认图纸专业和用途",
    )


def relative_label(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def classify(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in {".dwg", ".dxf", ".dwf", ".dwfx", ".pdf"}:
        return "drawing"
    if ext in {".xlsx", ".xls", ".csv", ".tsv"}:
        return "equipment table/BOM/inventory/price table"
    if ext in {".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff", ".bmp"}:
        return "screenshot"
    if ext in {".doc", ".docx", ".md", ".txt"}:
        return "standard/log"
    return mimetypes.guess_type(path.name)[0] or "unknown"


def is_hvac_relevant_path(path: Path) -> bool:
    text = str(path)
    return any(keyword in text for keyword in ["暖通", "空调", "多联机", "通风", "防排烟", "冷媒", "冷凝"])


def resolve_input_files(input_args: list[str], purpose: str) -> list[Path]:
    files: list[Path] = []
    for item in input_args:
        candidate = Path(item)
        path = candidate.resolve() if candidate.is_absolute() else (ROOT / candidate).resolve()
        if path.is_dir():
            files.extend(sorted(p.resolve() for p in path.rglob("*") if p.is_file() and not is_ignored_source(p)))
        elif path.is_file():
            if not is_ignored_source(path):
                files.append(path)
        else:
            raise FileNotFoundError(f"输入资料不存在：{item}")

    if purpose == "non_project_fixture":
        fixture_root = FIXTURE_ROOT.resolve()
        for path in files:
            if not str(path).startswith(str(fixture_root)):
                raise PermissionError(
                    "非项目验证模式只允许读取 tests/fixtures/ 下的夹具文件；"
                    f"拒绝读取：{path}"
                )
    return files


def is_ignored_source(path: Path) -> bool:
    name = path.name.lower()
    return path.suffix.lower() in {".dwl", ".dwl2", ".tmp"} or name.startswith("~$")


def inspect_workbook(source_id: str, path: Path) -> tuple[list[WorkbookSheetRow], str]:
    rows: list[WorkbookSheetRow] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
        for ws in wb.worksheets:
            used_max_row = 0
            used_max_column = 0
            for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
                for column_index, value in enumerate(row, start=1):
                    if value is not None and str(value).strip() != "":
                        used_max_row = max(used_max_row, row_index)
                        used_max_column = max(used_max_column, column_index)
            rows.append(
                WorkbookSheetRow(
                    source_id=source_id,
                    file_or_link=relative_label(path),
                    sheet_name=ws.title,
                    max_row=used_max_row,
                    max_column=used_max_column,
                )
            )
        wb.close()
    except Exception as exc:
        return rows, f"workbook inspect error: {exc}"
    summary = "; ".join(f"{row.sheet_name}({row.max_row}x{row.max_column})" for row in rows)
    return rows, summary or "no worksheets"


def value_at(row: tuple[object, ...], index: int) -> str:
    if index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def looks_like_item_code(value: str) -> bool:
    return bool(re.match(r"^[A-Za-z][A-Za-z0-9-]*$", value))


def extract_material_rows(source_id: str, path: Path) -> list[MaterialRow]:
    rows: list[MaterialRow] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return rows

    for ws in wb.worksheets:
        sheet_name = ws.title
        current_category = ""
        if sheet_name == "硬装材料":
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_number <= 5:
                    continue
                first = value_at(row, 0)
                item_name = value_at(row, 2)
                if first and not item_name and not value_at(row, 3) and not value_at(row, 7):
                    current_category = first
                    continue
                if not looks_like_item_code(first) or not item_name:
                    continue
                rows.append(
                    MaterialRow(
                        source_id=source_id,
                        file_or_link=relative_label(path),
                        sheet_name=sheet_name,
                        row_number=row_number,
                        item_code=first,
                        category=current_category,
                        item_name=item_name,
                        dimension_or_param=value_at(row, 3),
                        material=value_at(row, 4),
                        specification=value_at(row, 5),
                        installation_location=value_at(row, 7),
                        notes=value_at(row, 8),
                    )
                )
        elif sheet_name == "灯具材料":
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_number <= 2:
                    continue
                code = value_at(row, 0)
                name = value_at(row, 1)
                if not looks_like_item_code(code) or not name:
                    continue
                rows.append(
                    MaterialRow(
                        source_id=source_id,
                        file_or_link=relative_label(path),
                        sheet_name=sheet_name,
                        row_number=row_number,
                        item_code=code,
                        category="灯具",
                        item_name=name,
                        dimension_or_param=value_at(row, 3),
                        material=value_at(row, 4),
                        specification=value_at(row, 2),
                        installation_location=value_at(row, 7),
                        notes=value_at(row, 8),
                    )
                )
    wb.close()
    return rows


def extract_vrf_workbook(
    source_id: str, path: Path
) -> tuple[list[VrfSystemConfigRow], list[VrfEquipmentSummaryRow], list[VrfInstallMaterialRow]]:
    system_rows: list[VrfSystemConfigRow] = []
    equipment_rows: list[VrfEquipmentSummaryRow] = []
    material_rows: list[VrfInstallMaterialRow] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=False)
    except Exception:
        return system_rows, equipment_rows, material_rows

    for ws in wb.worksheets:
        if ws.title.startswith("系统配置"):
            current_area = ""
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_number <= 3:
                    continue
                seq = value_at(row, 0)
                if not seq:
                    continue
                area = value_at(row, 1) or current_area
                if value_at(row, 1):
                    current_area = value_at(row, 1)
                model = value_at(row, 6)
                equipment_type = value_at(row, 5)
                if not model and not equipment_type:
                    continue
                system_rows.append(
                    VrfSystemConfigRow(
                        source_id=source_id,
                        file_or_link=relative_label(path),
                        sheet_name=ws.title,
                        row_number=row_number,
                        floor_or_area=area,
                        room_name=value_at(row, 2),
                        area_m2=value_at(row, 3),
                        unit_load_w_m2=value_at(row, 4),
                        equipment_type=equipment_type,
                        equipment_model=model,
                        capacity_kw=value_at(row, 7),
                        quantity=value_at(row, 8),
                        cooling_capacity_kw=value_at(row, 9),
                        connection_ratio=value_at(row, 10),
                        system_mark=value_at(row, 11),
                        evidence=f"{source_id}:{ws.title}!row{row_number}",
                    )
                )
        elif ws.title == "设备安装汇总表":
            current_name = ""
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_number <= 3:
                    continue
                if not value_at(row, 0):
                    continue
                name = value_at(row, 1) or current_name
                if value_at(row, 1):
                    current_name = value_at(row, 1)
                if not value_at(row, 2):
                    continue
                equipment_rows.append(
                    VrfEquipmentSummaryRow(
                        source_id=source_id,
                        file_or_link=relative_label(path),
                        sheet_name=ws.title,
                        row_number=row_number,
                        equipment_name=name,
                        equipment_model=value_at(row, 2),
                        quantity=value_at(row, 3),
                        unit=value_at(row, 4),
                        unit_price=value_at(row, 5),
                        total_price=value_at(row, 6),
                        evidence=f"{source_id}:{ws.title}!row{row_number}",
                    )
                )
        elif ws.title == "安装材料":
            current_item_name = ""
            current_brand = ""
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_number <= 2:
                    continue
                if not value_at(row, 1) and not value_at(row, 2):
                    continue
                if value_at(row, 1):
                    current_item_name = value_at(row, 1)
                if value_at(row, 3):
                    current_brand = value_at(row, 3)
                floor_parts = []
                for col_index, floor_name in [(9, "一层"), (10, "二层"), (11, "三层"), (12, "四层"), (13, "五层"), (14, "总计")]:
                    value = value_at(row, col_index)
                    if value:
                        floor_parts.append(f"{floor_name}={value}")
                material_rows.append(
                    VrfInstallMaterialRow(
                        source_id=source_id,
                        file_or_link=relative_label(path),
                        sheet_name=ws.title,
                        row_number=row_number,
                        item_name=value_at(row, 1) or current_item_name,
                        spec=value_at(row, 2),
                        brand=value_at(row, 3) or current_brand,
                        quantity=value_at(row, 4),
                        unit=value_at(row, 5),
                        unit_price=value_at(row, 6),
                        total_price=value_at(row, 7),
                        notes=value_at(row, 8),
                        floor_breakdown="; ".join(floor_parts),
                        evidence=f"{source_id}:{ws.title}!row{row_number}",
                    )
                )
    wb.close()
    return system_rows, equipment_rows, material_rows


def compare_material_rows(material_rows: list[MaterialRow]) -> list[MaterialCompareRow]:
    grouped: dict[tuple[str, str], list[MaterialRow]] = {}
    for row in material_rows:
        grouped.setdefault((row.sheet_name, row.item_code), []).append(row)

    compare_rows: list[MaterialCompareRow] = []
    for (sheet_name, item_code), rows in sorted(grouped.items()):
        source_ids = sorted({row.source_id for row in rows})
        names = sorted({row.item_name for row in rows if row.item_name})
        dims = sorted({row.dimension_or_param for row in rows if row.dimension_or_param})
        locations = sorted({row.installation_location for row in rows if row.installation_location})
        same_count = len(source_ids) > 1
        same_name = len(names) <= 1
        same_dim = len(dims) <= 1
        if same_count and same_name and same_dim:
            conclusion = "same code in compared sheets"
        elif same_count:
            conclusion = "same code with differences"
        else:
            conclusion = "only in one source"
        evidence = "; ".join(f"{row.source_id}:{row.sheet_name}!row{row.row_number}" for row in rows)
        compare_rows.append(
            MaterialCompareRow(
                sheet_name=sheet_name,
                item_code=item_code,
                source_ids=", ".join(source_ids),
                item_names=" | ".join(names),
                dimensions_or_params=" | ".join(dims),
                installation_locations=" | ".join(locations),
                conclusion=conclusion,
                evidence=evidence,
            )
        )
    return compare_rows


def oda_installed() -> bool:
    return bool(
        shutil.which("ODAFileConverter.exe")
        or Path(r"C:\Program Files\ODA\ODAFileConverter\ODAFileConverter.exe").exists()
    )


def extract_dwfx(source_id: str, path: Path, extract_dir: Path) -> tuple[list[str], list[str], list[str]]:
    target = extract_dir / source_id
    if target.exists():
        shutil.rmtree(target)
    target.mkdir(parents=True, exist_ok=True)

    extracted: list[str] = []
    xml_summaries: list[str] = []
    image_summaries: list[str] = []
    with zipfile.ZipFile(path) as zf:
        for member in zf.infolist():
            if member.is_dir():
                continue
            out_path = target / member.filename
            out_path.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(member) as src, out_path.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            extracted.append(member.filename)

            suffix = out_path.suffix.lower()
            if suffix in {".xml", ".rels", ".fdseq", ".fdoc", ".fpage", ".dwfseq"}:
                try:
                    root = ET.parse(out_path).getroot()
                    xml_summaries.append(f"{member.filename}: root={root.tag.split('}')[-1]}")
                except ET.ParseError as exc:
                    xml_summaries.append(f"{member.filename}: XML parse error {exc}")
            elif suffix in {".png", ".jpg", ".jpeg"}:
                try:
                    with Image.open(out_path) as img:
                        image_summaries.append(f"{member.filename}: {img.width}x{img.height} {img.mode}")
                except Exception as exc:
                    image_summaries.append(f"{member.filename}: image read error {exc}")
    return extracted, xml_summaries, image_summaries


def write_csv(path: Path, rows: list[object]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(asdict(rows[0]).keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def add_sheet(wb: Workbook, title: str, rows: list[object]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["无数据"])
        return
    headers = list(asdict(rows[0]).keys())
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    for row in rows:
        ws.append([asdict(row)[h] for h in headers])
    ws.freeze_panes = "A2"
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 60)
        ws.column_dimensions[col[0].column_letter].width = width


def build_issue(issue_no: int, **kwargs: str) -> IssueRow:
    return IssueRow(issue_id=f"QA-{issue_no:03d}", **kwargs)


def build_outputs(
    output_dir: Path,
    purpose: str,
    source_rows: list[SourceRow],
    automation_rows: list[AutomationRow],
    issue_rows: list[IssueRow],
    dwfx_details: dict[str, dict[str, list[str]]],
    workbook_sheet_rows: list[WorkbookSheetRow],
    material_rows: list[MaterialRow],
    dwg_index_rows: list[DwgIndexRow],
    drawing_discipline_rows: list[DrawingDisciplineRow],
    material_compare_rows: list[MaterialCompareRow],
    vrf_system_rows: list[VrfSystemConfigRow],
    vrf_equipment_rows: list[VrfEquipmentSummaryRow],
    vrf_install_material_rows: list[VrfInstallMaterialRow],
) -> None:
    write_csv(output_dir / "source_register.csv", source_rows)
    write_csv(output_dir / "automation_assessment.csv", automation_rows)
    write_csv(output_dir / "issue_ledger.csv", issue_rows)
    write_csv(output_dir / "excel_workbook_index.csv", workbook_sheet_rows)
    write_csv(output_dir / "excel_material_rows.csv", material_rows)
    write_csv(output_dir / "dwg_version_index.csv", dwg_index_rows)
    write_csv(output_dir / "drawing_discipline_index.csv", drawing_discipline_rows)
    write_csv(output_dir / "excel_material_compare.csv", material_compare_rows)
    write_csv(output_dir / "vrf_system_config.csv", vrf_system_rows)
    write_csv(output_dir / "vrf_equipment_summary.csv", vrf_equipment_rows)
    write_csv(output_dir / "vrf_install_materials.csv", vrf_install_material_rows)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "资料登记", source_rows)
    add_sheet(wb, "自动化评估", automation_rows)
    add_sheet(wb, "问题台账", issue_rows)
    add_sheet(wb, "Excel工作簿索引", workbook_sheet_rows)
    add_sheet(wb, "Excel材料预览", material_rows)
    add_sheet(wb, "DWG版本索引", dwg_index_rows)
    add_sheet(wb, "图纸专业识别", drawing_discipline_rows)
    add_sheet(wb, "Excel材料对比", material_compare_rows)
    add_sheet(wb, "VRF系统配置", vrf_system_rows)
    add_sheet(wb, "VRF设备汇总", vrf_equipment_rows)
    add_sheet(wb, "VRF安装材料", vrf_install_material_rows)
    wb.save(output_dir / "workflow_pilot.xlsx")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        "# 多联机项目资料自动化工作流第一步报告",
        "",
        f"- 生成时间：{now}",
        f"- 工作目录：`{ROOT}`",
        f"- 运行用途：`{purpose}`",
        "- 入口规则：脚本不会默认扫描项目目录或桌面；必须通过 `--input` 显式指定资料。",
        "- 原始资料处理原则：只读扫描；派生清单、报告和解包内容均写入输出目录。",
        "",
        "## 当前结论",
        "",
        "- 第一阶段已形成可复用入口：资料登记、文件哈希、自动化可读性评估、问题台账、Excel 工作簿索引和 Excel/CSV/Markdown 交付物。",
        f"- 本次输入 DWG：{len(dwg_index_rows)} 个；Excel 材料预览：{len(material_rows)} 条；材料编码对比：{len(material_compare_rows)} 条。",
        f"- 暖通/空调相关 DWG：{sum(1 for row in drawing_discipline_rows if row.hvac_relevant == 'yes')} 个；专业交叉核对相关 DWG：{sum(1 for row in drawing_discipline_rows if row.hvac_relevant == 'partial')} 个。",
        f"- 暖通表格抽取：系统配置 {len(vrf_system_rows)} 行，设备汇总 {len(vrf_equipment_rows)} 行，安装材料 {len(vrf_install_material_rows)} 行。",
        "- 非项目验证模式只允许读取 `tests/fixtures/`；正式项目模式只处理用户显式指定的输入资料。",
        "- 设计、造价、审计结论不会在缺少正式任务范围、品牌、设备表、库存快照和报价口径时下定论。",
        "",
        "## 本机工具状态",
        "",
        "- Python/openpyxl/Pillow：已可用，用于 Excel、CSV、图片和报告生成。",
        f"- ezdxf：{'已安装' if importlib.util.find_spec('ezdxf') else '未安装'}，可用于 DXF 图元读取。",
        f"- ODA File Converter：{'已安装' if oda_installed() else '未安装'}，DWG 转 DXF 自动转换链路尚未完全打通。",
        "- 官方工具依据：ODA File Converter 官方说明支持 DWG/DXF 命令行批量转换；ezdxf 官方文档支持调用已安装的 ODA File Converter 读取 DWG。",
        "",
        "## 资料登记摘要",
        "",
        "| 编号 | 类型 | 文件 | 状态 | 说明 |",
        "| --- | --- | --- | --- | --- |",
    ]
    for row in source_rows:
        lines.append(f"| {row.source_id} | {row.type} | `{row.file_or_link}` | {row.status} | {row.notes} |")

    lines.extend(["", "## DWFX 解包摘要", ""])
    if not dwfx_details:
        lines.append("- 本次没有处理 DWFX 文件。")
    for source_id, detail in dwfx_details.items():
        lines.append(f"### {source_id}")
        lines.append(f"- 文件数量：{len(detail['files'])}")
        lines.append(f"- XML/页面摘要：{len(detail['xml'])} 项")
        lines.append(f"- 图片摘要：{len(detail['images'])} 项")
        for item in detail["images"][:5]:
            lines.append(f"- 预览图：{item}")

    lines.extend(
        [
            "",
            "## 待确认事项",
            "",
            "- 正式项目输入资料路径、资料版本和交付范围。",
            "- 项目品牌、系列、设备表、施工任务单/BOM 是否为最终版本。",
            "- 飞书库存表 Excel/CSV 快照或 API 接入方式；必须记录导出时间。",
            "- 报价口径：损耗率、人工费、管理费、利润、税率、运输/吊装/辅材边界。",
            "",
            "## 输出文件",
            "",
            "- `source_register.csv` / `workflow_pilot.xlsx`：资料登记",
            "- `automation_assessment.csv` / `workflow_pilot.xlsx`：自动化路径评估",
            "- `issue_ledger.csv` / `workflow_pilot.xlsx`：问题台账",
            "- `excel_workbook_index.csv` / `workflow_pilot.xlsx`：Excel 工作簿结构索引",
            "- `excel_material_rows.csv` / `workflow_pilot.xlsx`：Excel 材料行预览",
            "- `dwg_version_index.csv` / `workflow_pilot.xlsx`：DWG 头部版本索引",
            "- `drawing_discipline_index.csv` / `workflow_pilot.xlsx`：图纸专业识别",
            "- `excel_material_compare.csv` / `workflow_pilot.xlsx`：Excel 材料编码对比",
            "- `vrf_system_config.csv` / `workflow_pilot.xlsx`：暖通系统配置抽取",
            "- `vrf_equipment_summary.csv` / `workflow_pilot.xlsx`：暖通设备汇总抽取",
            "- `vrf_install_materials.csv` / `workflow_pilot.xlsx`：暖通安装材料抽取",
            "- `workflow_pilot_report.md`：本报告",
        ]
    )
    (output_dir / "workflow_pilot_report.md").write_text("\n".join(lines), encoding="utf-8-sig")


def run(args: argparse.Namespace) -> Path:
    files = sorted(resolve_input_files(args.input, args.purpose), key=lambda p: str(p).lower())

    output_dir = (ROOT / args.output_dir).resolve() if args.output_dir else DEFAULT_OUTPUT_DIR
    extract_dir = output_dir / "extracted"
    if output_dir.exists() and args.clean_output:
        resolved_outputs = (ROOT / "outputs").resolve()
        if not str(output_dir).startswith(str(resolved_outputs)):
            raise PermissionError(f"拒绝清理 outputs 目录之外的路径：{output_dir}")
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    extract_dir.mkdir(parents=True, exist_ok=True)

    source_rows: list[SourceRow] = []
    automation_rows: list[AutomationRow] = []
    issue_rows: list[IssueRow] = []
    dwfx_details: dict[str, dict[str, list[str]]] = {}
    workbook_sheet_rows: list[WorkbookSheetRow] = []
    material_rows: list[MaterialRow] = []
    dwg_index_rows: list[DwgIndexRow] = []
    drawing_discipline_rows: list[DrawingDisciplineRow] = []
    vrf_system_rows: list[VrfSystemConfigRow] = []
    vrf_equipment_rows: list[VrfEquipmentSummaryRow] = []
    vrf_install_material_rows: list[VrfInstallMaterialRow] = []

    for idx, path in enumerate(files, start=1):
        source_id = f"SRC-{idx:03d}"
        ext = path.suffix.lower()
        stat = path.stat()
        source_rows.append(
            SourceRow(
                source_id=source_id,
                file_or_link=relative_label(path),
                type=classify(path),
                version_date=datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d"),
                owner_role="commander",
                status="used" if is_hvac_relevant_path(path) else "registered_only",
                notes=f"大小 {stat.st_size} bytes；修改时间 {datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')}",
                sha256=sha256(path),
            )
        )

        if ext == ".dwfx":
            extracted, xml_summaries, image_summaries = extract_dwfx(source_id, path, extract_dir)
            dwfx_details[source_id] = {"files": extracted, "xml": xml_summaries, "images": image_summaries}
            automation_rows.append(
                AutomationRow(
                    source_id=source_id,
                    file_or_link=relative_label(path),
                    format=ext,
                    readable_now="yes",
                    extraction_route="zipfile 解包 -> XML/FixedPage/PNG 结构索引 -> 人工复核页面内容",
                    extracted_artifacts=f"{len(extracted)} files; {len(xml_summaries)} XML-like; {len(image_summaries)} images",
                    confidence="medium",
                    manual_review_point="需人工确认图纸比例、图层语义、设备表位置和数量口径",
                    next_tool_need="DWG/DXF 原图或官方转换器，用于图元级抽取",
                )
            )
        elif ext == ".dwg":
            dwg_index_rows.append(inspect_dwg_header(source_id, path))
            drawing_discipline_rows.append(classify_drawing_discipline(source_id, path))
            automation_rows.append(
                AutomationRow(
                    source_id=source_id,
                    file_or_link=relative_label(path),
                    format=ext,
                    readable_now="no" if not oda_installed() else "yes",
                    extraction_route="ODA File Converter -> DXF -> ezdxf 图元/文字/块属性读取",
                    extracted_artifacts="not extracted" if not oda_installed() else "ready for conversion",
                    confidence="low" if not oda_installed() else "medium",
                    manual_review_point="需确认 CAD 版本、图层命名、比例、块属性和专业范围",
                    next_tool_need="安装/验证 ODA File Converter；或由设计软件导出 DXF",
                )
            )
        elif ext == ".dxf":
            automation_rows.append(
                AutomationRow(
                    source_id=source_id,
                    file_or_link=relative_label(path),
                    format=ext,
                    readable_now="yes",
                    extraction_route="ezdxf -> 图元/文字/块属性读取 -> 人工复核",
                    extracted_artifacts="registered for DXF workflow",
                    confidence="medium",
                    manual_review_point="需人工确认图层语义、比例、单位和标注是否完整",
                    next_tool_need="无；进入 DXF 抽取脚本开发",
                )
            )
        elif ext in {".xlsx", ".xls", ".csv"}:
            sheet_summary = "csv file"
            if ext in {".xlsx", ".xls"}:
                workbook_rows, sheet_summary = inspect_workbook(source_id, path)
                workbook_sheet_rows.extend(workbook_rows)
                if is_hvac_relevant_path(path):
                    material_rows.extend(extract_material_rows(source_id, path))
                    system_rows, equipment_rows, install_material_rows = extract_vrf_workbook(source_id, path)
                    vrf_system_rows.extend(system_rows)
                    vrf_equipment_rows.extend(equipment_rows)
                    vrf_install_material_rows.extend(install_material_rows)
            automation_rows.append(
                AutomationRow(
                    source_id=source_id,
                    file_or_link=relative_label(path),
                    format=ext,
                    readable_now="yes",
                    extraction_route="openpyxl/csv -> 工作表/单元格/公式提取 -> 字段标准化" if is_hvac_relevant_path(path) else "非暖通资料：仅登记工作簿结构，不展开材料明细",
                    extracted_artifacts=sheet_summary,
                    confidence="high" if is_hvac_relevant_path(path) else "not applicable",
                    manual_review_point="需确认表头、版本、导出时间、单位和公式口径" if is_hvac_relevant_path(path) else "非暖通资料，不进入本团队展开工作",
                    next_tool_need="无；库存 API 未接入时使用 Excel/CSV 快照" if is_hvac_relevant_path(path) else "等待暖通设备表/BOM/库存快照",
                )
            )

    material_compare_rows = compare_material_rows(material_rows)

    issue_no = 1
    if not any(row.format in {".dwg", ".dxf"} for row in automation_rows):
        issue_rows.append(
            build_issue(
                issue_no,
                severity="S2",
                source="automation",
                description="本次输入没有 DWG/DXF，无法验证图元级工程量自动提取。",
                impact="CAD 图层、块属性、线段长度、设备编号无法直接结构化抽取",
                evidence="本次显式输入资料清单未包含 .dwg/.dxf",
                owner="automation-engineer",
                recommendation="正式项目阶段提供 DWG/DXF，或由设计软件导出 DXF 后再进行图元级验证",
                status="open",
                release_conclusion="conditional pass",
            )
        )
        issue_no += 1
    if not oda_installed():
        issue_rows.append(
            build_issue(
                issue_no,
                severity="S2",
                source="automation",
                description="本机未安装 ODA File Converter，DWG 转 DXF 自动链路尚未打通。",
                impact="拿到 DWG 后暂不能直接完成命令行批量转换和 ezdxf 图元读取验证",
                evidence="未找到 ODAFileConverter.exe；默认路径 C:\\Program Files\\ODA\\ODAFileConverter\\ODAFileConverter.exe 不存在",
                owner="automation-engineer",
                recommendation="通过 ODA 官方下载/试用入口安装授权版本；安装后运行 DWG -> DXF -> ezdxf 抽取回归测试",
                status="open",
                release_conclusion="conditional pass",
            )
        )
        issue_no += 1
    if dwg_index_rows and not any(row.hvac_relevant == "yes" for row in drawing_discipline_rows):
        issue_rows.append(
            build_issue(
                issue_no,
                severity="S2",
                source="design",
                description="本次授权目录未识别到暖通/空调专业 DWG，暂不能开展多联机设计校核。",
                impact="不能校核室内外机、冷媒管、分歧管、管长、高差和制冷剂追加规则",
                evidence="drawing_discipline_index.csv 中未出现 hvac_relevant=yes",
                owner="designer",
                recommendation="补充暖通/空调专业施工图、设备表或设计说明；当前装饰/电气/给排水图只适合做交叉条件核对",
                status="open",
                release_conclusion="conditional pass",
            )
        )
        issue_no += 1
    if not any(row.format in {".xlsx", ".xls", ".csv"} and "非暖通资料" not in row.extraction_route for row in automation_rows):
        issue_rows.append(
            build_issue(
                issue_no,
                severity="S2",
                source="cost/quote",
                description="本次输入没有暖通相关 Excel/CSV 表格，无法验证设备表、BOM 或库存快照链路。",
                impact="不能验证暖通设备表、工程量核对、库存优先和报价基础表字段",
                evidence="本次显式输入虽包含 Excel，但未识别为暖通/空调相关资料",
                owner="cost-estimator",
                recommendation="正式项目阶段提供施工任务单、BOM、飞书库存导出或价格表快照",
                status="open",
                release_conclusion="conditional pass",
            )
        )

    build_outputs(
        output_dir,
        args.purpose,
        source_rows,
        automation_rows,
        issue_rows,
        dwfx_details,
        workbook_sheet_rows,
        material_rows,
        dwg_index_rows,
        drawing_discipline_rows,
        material_compare_rows,
        vrf_system_rows,
        vrf_equipment_rows,
        vrf_install_material_rows,
    )
    return output_dir


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="多联机/暖通项目资料自动化工作流第一步。")
    parser.add_argument("--input", action="append", required=True, help="显式指定输入文件或目录，可重复传入。")
    parser.add_argument(
        "--purpose",
        required=True,
        choices=["formal_project", "non_project_fixture"],
        help="formal_project 表示用户指定的正式项目资料；non_project_fixture 只允许 tests/fixtures/。",
    )
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR.relative_to(ROOT)), help="输出目录，默认 outputs/auto_workflow_pilot。")
    parser.add_argument("--clean-output", action="store_true", help="运行前清理输出目录，仅允许清理 outputs/ 下路径。")
    return parser.parse_args()


def main() -> None:
    output_dir = run(parse_args())
    print(f"OK: {output_dir}")


if __name__ == "__main__":
    main()
